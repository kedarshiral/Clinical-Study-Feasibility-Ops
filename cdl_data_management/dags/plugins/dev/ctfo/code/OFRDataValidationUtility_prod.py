import boto3
import io
import pandas as pd
from datetime import datetime, date
import time
import traceback
import os
import sys
import copy
from openpyxl import load_workbook
import smtplib
import smtplib, ssl
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication
import pandasql as ps
import subprocess

service_directory_path = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(1, service_directory_path)
code_dir_path = os.path.abspath(os.path.join(service_directory_path, "../"))
sys.path.insert(1, code_dir_path)
from MySQLConnectionManager import MySQLConnectionManager
from NotificationUtility import NotificationUtility
from ExecutionContext import ExecutionContext
import CommonConstants
from CommonUtils import CommonUtils
from ConfigUtility import JsonConfigUtility
from LogSetup import logger

MODULE_NAME = "Validation_utility"
execution_context = ExecutionContext()
execution_context.set_context({"module_name": MODULE_NAME})
MODULE_NAME = 'OFRDataValidation'
sender_email = "sanofi.clinical@zs.com"
recipient_list = [
    "sankarshana.kadambari@zs.com",
    "dhananjay.nagare@zs.com",
    "vivek.pangasa@zs.com",
    "richa.d.singh@zs.com",
    "archit.s.aggarwal@zs.com"]


def make_dw_mail(dw_df, subject, process_id):
    msg = MIMEMultipart()
    msg['Subject'] = subject + " " + str(int(float(process_id)))
    html = """\
    <html>
      <head></head>
      <body>
        <p>Hi Team,<br>
        <br><br>
        Please find below run time details for Process ID : {0}
        <br><br>DW Run Time:<br>
           {1}
          <br><br>
           Regards,
           <br>
           DigitalHealth R&D Advanced Analytics Platform Infra Support
           <br>
           <br>
           <br>
           <br>
           <br>
           <br>
           </p>
           <p style="text-align:center;">
           This email is auto-generated. Do NOT Reply to this message</p>
      </body>
    </html>
    """.format(str(process_id), dw_df.to_html(index=False))
    partHTML = MIMEText(html, 'html')
    msg.attach(partHTML)
    return msg


def disease_mapping_validation_mail(res1, res2, res3, res4, res5, res6, res7, res8, d1, d2, d3, d4, d5, d6, d7, d8):
    html_mesage = """\
        <html>
          <head></head>
          <body>
            <p>Hi Team,<br>
            <br><br>
            Please find below Validation Details For Disease Mapping
            <br><br><b>1.{8}:</b><br>
               {0}
               <br><br><b>2. {9}:</b><br>
               {1}
               <br><br><b>3. {10}:</b><br>
               {2}
               <br><br><b>4. {11}:</b><br>
               {3}
              <br><br><b>5. {12}:</b><br>
               {4}
               <br><br><b>6. {13}:</b><br>
               {5}
               <br><br><b>7. {14}:</b><br>
               {6}
              <br><br><b>8. {15}:</b><br>
               {7}
              <br><br>
              If the count in the Result of Test Case 7 doesn't match, then You will have to read the Disease Mapping and execute the below query:
              <br><br>
              select distinct(a.standard_disease) from disease_map a left join disease_map b on lower(a.standard_disease) =lower(b.standard_disease)  where a.standard_disease <> b.standard_disease
              <br><br>
               Regards,
               <br>
               Sanofi DE Team
               <br>
               Clinical Development Excellence
               <br>
               <br>
               <br>
               <br>
               <br>
               </p>
               <p style="text-align:center;">
               This email is auto-generated. Do NOT Reply to this message</p>
          </body>
        </html>
        """.format(res1.to_html(index=False), res2.to_html(index=False),
                   res3.to_html(index=False), res4.to_html(index=False), res5.to_html(index=False),
                   res6.to_html(index=False), res7.to_html(index=False), res8.to_html(index=False), str(d1), str(d2),
                   str(d3), str(d4), str(d5), str(d6), str(d7), str(d8))
    return html_mesage


def ta_mapping_validation_mail(res1, res2, res3, res4, res5, res6, res7, res8, res9, res10, d1, d2, d3, d4, d5, d6, d7,
                               d8, d9, d10):
    html_mesage = """\
        <html>
          <head></head>
          <body>
            <p>Hi Team,<br>
            <br><br>
            Please find below Validation Details For TA Mapping
            <br><br><b>1.{10}:</b><br>
               {0}
               <br><br><b>2.{11}:</b><br>
               {1}
               <br><br><b>3.{12}:</b><br>
               {2}
               <br><br><b>4.{13}:</b><br>
               {3}
               <br><br><b>5.{14}:</b><br>
               {4}
               <br><br><b>6.{15}:</b><br>
               {5}
               <br><br><b>7.{16}:</b><br>
               {6}
               <br><br><b>8.{17}:</b><br>
               {7}
               <br><br><b>9.{18}:</b><br>
               {8}
               <br><br><b>10.{19}:</b><br>
               {9}
              <br><br>
              If the count in the Result of Test Case 8 doesn't match, then You will have to read the Ta Mapping and execute the below query:
              <br><br>
              select distinct(a.Disease_name) from ta_map a left join ta_map b on lower(a.therapeutic_area) =lower(b.therapeutic_area)  where a.therapeutic_area <> b.therapeutic_area
              <br><br>
               Regards,
               <br>
               Sanofi DE Team
               <br>
               Clinical Development Excellence
               <br>
               <br>
               <br>
               <br>
               <br>
               </p>
               <p style="text-align:center;">
               This email is auto-generated. Do NOT Reply to this message</p>
          </body>
        </html>
        """.format(res1.to_html(index=False), res2.to_html(index=False),
                   res3.to_html(index=False), res4.to_html(index=False), res5.to_html(index=False),
                   res6.to_html(index=False), res7.to_html(index=False), res8.to_html(index=False),
                   res9.to_html(index=False), res10.to_html(index=False), str(d1), str(d2),
                   str(d3), str(d4), str(d5), str(d6), str(d7), str(d8), str(d9), str(d10))
    return html_mesage


def make_dw_validation_mail(subject):
    html_mesage = """\
        <html>
          <head></head>
          <body>
            <p>Hi Team,<br>
            <br><br>
            Please find below Validation Details For DW Dag
            <br><br>Athena Check:<br>
               {0}
              <br><br>
               Regards,
               <br>
               Sanofi DE Team
               <br>
               Clinical Development Excellence
               <br>
               <br>
               <br>
               <br>
               <br>
               </p>
               <p style="text-align:center;">
               This email is auto-generated. Do NOT Reply to this message</p>
          </body>
        </html>
        """.format(subject.to_html(index=False))
    return html_mesage


def make_ingestion_mail(adapter_df, ingestion_df, file_count_df, dqm_df, athena_df, process_id):
    html_mesage = """\
    <html>
      <head></head>
      <body>
        <p>Hi Team,<br>
        <br><br>
        Please find below RDS Validation Details For Process ID : {0}
        <br><br>Adapter Run Time:<br>
           {1}
           <br><br>Ingestion Run Time:<br>
           {2}
           <br><br>File Count Check:<br>
           {3}
           <br><br>DQM Check:<br>
           {4}
          <br><br>Athena Check:<br>
           {5}
          <br><br>
           Regards,
           <br>
           Sanofi DE Team
           <br>
           Clinical Development Excellence
           <br>
           <br>
           <br>
           <br>
           <br>
           </p>
           <p style="text-align:center;">
           This email is auto-generated. Do NOT Reply to this message</p>
      </body>
    </html>
    """.format(str(process_id), adapter_df.to_html(index=False), ingestion_df.to_html(index=False),
               file_count_df.to_html(index=False), dqm_df.to_html(index=False), athena_df.to_html(index=False))
    return html_mesage


def make_kpi_ops_mail(subject):
    html_mesage = """\
        <html>
          <head></head>
          <body>
            <p>Hi Team,<br>
            <br><br>
            Please find below KPI Validation Details For This week
            <br><br>Athena Check:<br>
               {0}
              <br><br>
               Regards,
               <br>
               Sanofi DE Team
               <br>
               Clinical Development Excellence
               <br>
               <br>
               <br>
               <br>
               <br>
               </p>
               <p style="text-align:center;">
               This email is auto-generated. Do NOT Reply to this message</p>
          </body>
        </html>
        """.format(subject.to_html(index=False))
    return html_mesage


class OFRDataValidationUtility(object):
    def __init__(self):
        self.execution_context = ExecutionContext()
        self.execution_context.set_context({"module_name": MODULE_NAME})
        self.configuration = JsonConfigUtility(os.path.join(CommonConstants.AIRFLOW_CODE_PATH,
                                                            CommonConstants.ENVIRONMENT_CONFIG_FILE))
        self.region = self.configuration.get_configuration([CommonConstants.ENVIRONMENT_PARAMS_KEY, "s3_region"])
        self.audit_db = self.configuration.get_configuration([CommonConstants.ENVIRONMENT_PARAMS_KEY, "mysql_db"])
        self.s3_sandbox_bucket = self.configuration.get_configuration(
            [CommonConstants.ENVIRONMENT_PARAMS_KEY, "s3_sandbox_bucket"])
        self.cycle_status_succeed = CommonConstants.STATUS_SUCCEEDED
        self.query_loc = CommonConstants.RDS_QUERY_PATH
        self.query_path = CommonConstants.OFR_QUERY_PATH
        # self.qc_query_path = CommonConstants.QC_QUERY_PATH
        self.s3 = boto3.client('s3', region_name=self.region)
        self.query_obj = self.s3.get_object(Bucket=self.s3_sandbox_bucket, Key=self.query_loc)
        self.adapter_obj = self.s3.get_object(Bucket=self.s3_sandbox_bucket, Key=self.query_loc)
        self.ofr_obj = self.s3.get_object(Bucket=self.s3_sandbox_bucket, Key=self.query_path)
        # self.qc_obj = self.s3.get_object(Bucket=self.s3_sandbox_bucket, Key=self.qc_query_path)
        self.validation_queries = pd.read_excel(io.BytesIO(self.query_obj['Body'].read()), "queries")
        self.adapter_mapping = pd.read_excel(io.BytesIO(self.adapter_obj['Body'].read()), "adapter_mapping")
        self.ec2_excel_path = CommonConstants.RUN_TIME_OUTPUT_LOCAL_PATH
        self.run_time_path = CommonConstants.RUN_TIME_OUTPUT_PATH
        self.create_sheet_obj = self.s3.get_object(Bucket=self.s3_sandbox_bucket, Key=self.run_time_path)
        self.run_time_obj = self.s3.get_object(Bucket=self.s3_sandbox_bucket, Key=self.query_loc)
        self.run_time_df = pd.read_excel(io.BytesIO(self.run_time_obj['Body'].read()), "process_mapping")
        self.msck_repair_tc_key = CommonConstants.MSCK_REPAIR_TC_KEY
        self.ofr_validation_queries = pd.read_excel(io.BytesIO(self.ofr_obj['Body'].read()))
        # self.qc_validation_queries = pd.read_excel(io.BytesIO(self.qc_obj['Body'].read()))
        self.ofr_validation_output_file_path = CommonConstants.OFR_VALIDATION_OUTPUT_FILE_PATH
        self.ofr_validation_output_file_s3_path = CommonConstants.OFR_VALIDATION_OUTPUT_FILE_S3_PATH
        self.rds_email_type = 'ofr_rds_data_validation'
        self.ofr_dw_email_type = 'ofr_dw_data_validation_info'
        self.ofr_validation_output_file_name = CommonConstants.OFR_VALIDATION_OUTPUT_FILE_NAME

    def adapter_run_time(self, process_id):
        try:
            start_time = ""
            process_name = ""
            l2_run_time_hours = ""
            process_id_list = []
            process_name_list = []
            data_date_list = []
            l2_run_time_hours_list = []
            query = self.validation_queries[self.validation_queries["query_type"] == "adapter_run_time"].iloc[0][
                'query']
            adapter_df = self.adapter_mapping[self.adapter_mapping["process_id"].astype(str) == str(process_id)]
            mysql_obj = MySQLConnectionManager()
            for index, row in adapter_df.iterrows():
                source = row["source_name"]
                status = row["status"]
                final_query = query.replace("$$status", status).replace("$$source_name", source)
                status_msg = "The Query to Execute for Adapter RUN time with process id " + str(
                    process_id) + ": " + final_query
                logger.info(status_msg, extra=self.execution_context.get_context())
                response = mysql_obj.execute_query_mysql(final_query)
                response_list = list(response)
                process_id_list.append(str(process_id))
                data_date = ""
                process_name = ""
                l2_run_time_hours = ""
                response_df = pd.DataFrame(response_list)
                for i, r in response_df.iterrows():
                    process_name = r["process_name"]
                    data_date = r["data_date"]
                    l2_run_time_hours = r["L2_run_time_in_hours"]
                    if process_name and data_date and l2_run_time_hours:
                        self.adapter_update_run_time(str(process_name), str(data_date), str(l2_run_time_hours))
                    else:
                        pass
                process_name_list.append(process_name)
                data_date_list.append(str(data_date))
                l2_run_time_hours_list.append(str(l2_run_time_hours))
            list_of_tuples = list(zip(process_id_list, process_name_list, data_date_list, l2_run_time_hours_list))
            df_final = pd.DataFrame(list_of_tuples,
                                    columns=["process_id", "adapter_name", "data_date", "L0_run_time_in_hours"])
            logger.info(df_final, extra=self.execution_context.get_context())
            if len(df_final) == 0:
                process_id_list_str = [str(process_id)]
                response_list = ["No Records Found"]
                list_of_tuples = list(zip(process_id_list_str, response_list))
                res_df = pd.DataFrame(list_of_tuples,
                                      columns=["process_id", "adapter_run_time"])
            else:
                res_df = df_final
            return res_df
        except Exception as e:
            error = "ERROR in " + "Adapter Run Time" + \
                    "For Proceess ID : " + str(process_id) + \
                    " ERROR MESSAGE: " + str(e)
            self.execution_context.set_context({"traceback": error})
            logger.error(error, extra=self.execution_context.get_context())

    def ingestion_run_time(self, process_id):
        try:
            query = self.validation_queries[self.validation_queries["query_type"] == "ingestion_run_time"].iloc[0][
                'query']
            final_query = query.replace("$$process_id", str(process_id))
            status_msg = "The Query to Execute for Ingestion RUN time with process id " + str(
                process_id) + ": " + final_query
            logger.info(status_msg, extra=self.execution_context.get_context())
            mysql_obj = MySQLConnectionManager()
            response = mysql_obj.execute_query_mysql(final_query)
            if response:
                response_list = list(response)
                response_df = pd.DataFrame(response_list)
                res_df = response_df[["process_id", "process_name", "data_date", "L1_run_time_in_hours"]]
                logger.info(res_df, extra=self.execution_context.get_context())
                send_data_date = str(res_df["data_date"].to_string(index=False))
                run_time = str(res_df["L1_run_time_in_hours"].to_string(index=False))
                self.update_run_time(process_id, send_data_date, run_time)
            else:
                process_id_list_str = [str(process_id)]
                response_list = ["No Records Found"]
                list_of_tuples = list(zip(process_id_list_str, response_list))
                res_df = pd.DataFrame(list_of_tuples,
                                      columns=["process_id", "ingestion_run_time"])
            return res_df
        except Exception as e:
            error = "ERROR in " + "Ingestion Run Time" + \
                    "For Proceess ID : " + str(process_id) + \
                    " ERROR MESSAGE: " + str(e)
            self.execution_context.set_context({"traceback": error})
            logger.error(error, extra=self.execution_context.get_context())

    def dw_run_time(self, process_id, **kwargs):
        try:
            query = self.validation_queries[self.validation_queries["query_type"] == "dw_run_time"].iloc[0]['query']
            final_query = query.replace("$$process_id", str(process_id))
            status_msg = "The Query to Execute for DW RUN time with process id " + str(process_id) + ": " + final_query
            logger.info(status_msg, extra=self.execution_context.get_context())
            mysql_obj = MySQLConnectionManager()
            response = mysql_obj.execute_query_mysql(final_query)
            if response:
                response_list = list(response)
                response_df = pd.DataFrame(response_list)
                res_df = response_df[["process_id", "process_name", "data_date", "L2_run_time_in_hours"]]
                logger.info(res_df, extra=self.execution_context.get_context())
                send_data_date = str(res_df["data_date"].to_string(index=False))
                run_time = str(res_df["L2_run_time_in_hours"].to_string(index=False))
                self.update_run_time(process_id, send_data_date, run_time)
            else:
                process_id_list_str = [str(process_id)]
                response_list = ["No Records Found"]
                list_of_tuples = list(zip(process_id_list_str, response_list))
                res_df = pd.DataFrame(list_of_tuples,
                                      columns=["process_id", "dw_run_time"])
            email_type = "ofr_rds_data_validation"
            ses = self.configuration.get_configuration(
                [CommonConstants.ENVIRONMENT_PARAMS_KEY, "ses_region"])
            logger.info("SES Region retrieved from Configurations is: " + str(ses),
                        extra=self.execution_context.get_context())
            email_sender = self.configuration.get_configuration([CommonConstants.ENVIRONMENT_PARAMS_KEY,
                                                                 "email_type_configurations",
                                                                 str(email_type),
                                                                 "ses_sender"])
            logger.info("Email Sender retrieved from Configurations is: " + str(email_sender),
                        extra=self.execution_context.get_context())
            email_recipient_list = self.configuration.get_configuration(
                [CommonConstants.ENVIRONMENT_PARAMS_KEY, "email_type_configurations",
                 str(email_type), "ses_recipient_list"])
            logger.info("Email Recipient retrieved from Configurations is: " + str(
                email_recipient_list), extra=self.execution_context.get_context())
            email_subject = self.configuration.get_configuration([CommonConstants.ENVIRONMENT_PARAMS_KEY,
                                                                  "email_type_configurations",
                                                                  str(email_type),
                                                                  "subject_dw"])
            env = self.configuration.get_configuration([CommonConstants.ENVIRONMENT_PARAMS_KEY, "environment"])
            email_subject = "[" + str(env).upper() + "] " + email_subject
            msg = make_dw_mail(res_df, email_subject, process_id)
            # And finally, send the email
            print("Email Sender retrieved from Configurations is: " + str(email_sender))
            email_recipient_list = self.configuration.get_configuration(
                [CommonConstants.ENVIRONMENT_PARAMS_KEY, "delta_email_type_configurations", "ses_recipient_list"])
            sender = "sanofi.clinical@zs.com"
            status = self.send_email(sender, email_recipient_list, str(msg), email_subject)
            if status:
                print("Email sent Successfully")
            else:
                print("Check Logs")
            # ses = boto3.client("ses", region_name=self.region)
            # ses.send_raw_email(
            #     Source=email_sender,
            #     Destinations=email_recipient_list,
            #     RawMessage={
            #         'Data': msg.as_string(),
            #     }
            # )
            return res_df
        except Exception as e:
            error = "ERROR in " + "DW Run Time" + \
                    "For Proceess ID : " + str(process_id) + \
                    " ERROR MESSAGE: " + str(e)
            self.execution_context.set_context({"traceback": error})
            logger.error(error, extra=self.execution_context.get_context())

    def file_count_check(self, process_id):
        try:
            file_success_list = []
            count = 0
            query = self.validation_queries[self.validation_queries["query_type"] == "file_count_check"].iloc[0][
                'query']
            final_query = query.replace("$$process_id", str(process_id))
            status_msg = "The Query to Execute for file count Check with process id " + str(
                process_id) + ": " + final_query
            logger.info(status_msg, extra=self.execution_context.get_context())
            mysql_obj = MySQLConnectionManager()
            response = mysql_obj.execute_query_mysql(final_query)
            if response:
                response_list = list(response)
                response_df = pd.DataFrame(response_list)
                for i, r in response_df.iterrows():
                    if str(r["file_status"]).lower() == "succeeded":
                        count = count + 1
                    else:
                        pass
                file_success_list.append(str(count))
                total_file = str(len(response_df))
                start_time = str(response_df["file_process_start_time"][0])
                query_to_execute = "select distinct(workflow_name) from " + self.audit_db + "." + CommonConstants.EMR_PROCESS_WORKFLOW_MAP_TABLE + \
                                   " where process_id = $$process_id;"
                query_result = mysql_obj.execute_query_mysql(query_to_execute.replace("$$process_id", str(process_id)))
                query_result_list = list(query_result)
                query_result_df = pd.DataFrame(query_result_list)
                source = str(query_result_df["workflow_name"].to_string(index=False))
                source_list = [source]
                total_file_list = [total_file]
                process_id_list = [str(process_id)]
                list_of_tuples = list(zip(process_id_list, source_list, total_file_list, file_success_list))
                res_df = pd.DataFrame(list_of_tuples,
                                      columns=["process_id", "process_name", "total_file", "file_success_count"])
                logger.info(res_df, extra=self.execution_context.get_context())

            else:
                process_id_list_str = [str(process_id)]
                response_list = ["No Records Found"]
                list_of_tuples = list(zip(process_id_list_str, response_list))
                res_df = pd.DataFrame(list_of_tuples,
                                      columns=["process_id", "file_count_check"])
            return res_df
        except Exception as e:
            error = "ERROR in " + "File Count Check" + \
                    "For Proceess ID : " + str(process_id) + \
                    " ERROR MESSAGE: " + str(e)
            self.execution_context.set_context({"traceback": error})
            logger.error(error, extra=self.execution_context.get_context())

    def dqm_check(self, process_id):
        try:
            mysql_obj = MySQLConnectionManager()
            query = self.validation_queries[self.validation_queries["query_type"] == "dqm_check"].iloc[0]['query']
            final_query = query.replace("$$process_id", str(process_id))
            status_msg = "The Query to Execute for DQM Check with process id " + str(process_id) + ": " + final_query
            logger.info(status_msg, extra=self.execution_context.get_context())
            response = mysql_obj.execute_query_mysql(final_query)
            if response:
                response_list = list(response)
                response_df = pd.DataFrame(response_list)
                response_df = response_df[
                    ["batch_id", "column_name", "qc_type", "criticality", "dataset_id", "qc_failure_percentage",
                     "qc_id", "qc_message", "qc_status"]]
                logger.info(response_df, extra=self.execution_context.get_context())
                if len(response_df) == 0:
                    process_id_list_str = [str(process_id)]
                    response_list = ["No records found failing critical check"]
                    list_of_tuples = list(zip(process_id_list_str, response_list))
                    res_df = pd.DataFrame(list_of_tuples,
                                          columns=["process_id", "qc_message"])
                else:
                    res_df = response_df
            else:
                process_id_list_str = [str(process_id)]
                response_list = ["No Records Found Failing Critical check"]
                list_of_tuples = list(zip(process_id_list_str, response_list))
                res_df = pd.DataFrame(list_of_tuples,
                                      columns=["process_id", "qc_message"])
            return res_df
        except Exception as e:
            error = "ERROR in " + "DQM Check" + \
                    "For Proceess ID : " + str(process_id) + \
                    " ERROR MESSAGE: " + str(e)
            self.execution_context.set_context({"traceback": error})
            logger.error(error, extra=self.execution_context.get_context())

    def fetch_dataset_id(self, process_id):
        try:
            mysql_obj = MySQLConnectionManager()
            query = """select a.dataset_id, a.dataset_name, d.prev_batch_id,b.curr_batch_id from """ + self.audit_db + """.ctl_dataset_master a inner join
            (select max(batch_id) as curr_batch_id, dataset_id from """ + self.audit_db + """.log_file_smry where file_status = 'SUCCEEDED' and process_id = """ + str(
                process_id) + """ group by 2) b
            on b.dataset_id = a.dataset_id
            inner join
            (select max(batch_id) as prev_batch_id,dataset_id from """ + self.audit_db + """.log_file_smry where file_status = 'SUCCEEDED' and batch_id in
            (select b_id from
            (select max(batch_id) as b_id, dataset_id from """ + self.audit_db + """.log_file_smry where process_id = """ + str(
                process_id) + """ and file_status = 'SUCCEEDED' and batch_id not in
            (select b_id from
            (select max(batch_id) as b_id, dataset_id from """ + self.audit_db + """.log_file_smry where file_status = 'SUCCEEDED' and process_id = """ + str(
                process_id) + """ group by 2) a) group by 2) e ) group by 2) d
            on b.dataset_id = d.dataset_id;"""
            status_msg = "The Query to Execute for getting current and previous run batch ids for process_id " + \
                         str(process_id) + ": " + query
            logger.info(status_msg, extra=self.execution_context.get_context())
            execute_query = mysql_obj.execute_query_mysql(query)
            respnse_list = list(execute_query)
            batch_id_df = pd.DataFrame(respnse_list)
            logger.info(batch_id_df, extra=self.execution_context.get_context())
            return batch_id_df
        except:
            logger.error("Error occured while data validation", extra=self.execution_context.get_context())
            raise

    def query_results(self, query, key):
        session = boto3.Session()
        athena_results_path = self.configuration.get_configuration(
            [CommonConstants.ENVIRONMENT_PARAMS_KEY, "athena_results_path"])
        WorkgroupName = CommonConstants.WorkgroupName
        client = session.client('athena', region_name=self.region)
        configuration = JsonConfigUtility(os.path.join(CommonConstants.AIRFLOW_CODE_PATH,
                                                       CommonConstants.ENVIRONMENT_CONFIG_FILE))
        env = configuration.get_configuration([CommonConstants.ENVIRONMENT_PARAMS_KEY, "delta_automator_env"])
        WorkgroupName = WorkgroupName.replace("$$env", env)
        logger.debug(query)
        try:
            ## This function executes the query and returns the query execution ID
            response_query_execution_id = client.start_query_execution(
                QueryString=query,
                ResultConfiguration={
                    'OutputLocation': athena_results_path
                },
                WorkGroup=WorkgroupName
            )
            ## This function takes query execution id as input and returns the details of the query executed
            response_get_query_details = client.get_query_execution(
                QueryExecutionId=response_query_execution_id['QueryExecutionId']
            )
            # time.sleep(1)
            ## Condition for checking the details of response
            status = 'RUNNING'
            iterations = 100
            while (iterations > 0):
                iterations = iterations - 1
                response_get_query_details = client.get_query_execution(
                    QueryExecutionId=response_query_execution_id['QueryExecutionId']
                )
                status = response_get_query_details['QueryExecution']['Status']['State']
                logger.debug(status, extra=self.execution_context.get_context())
                if (status == 'FAILED') or (status == 'CANCELLED'):
                    return False
                elif status == 'SUCCEEDED':
                    location = response_get_query_details['QueryExecution']['ResultConfiguration']['OutputLocation']
                    ## Function to get output results
                    response_query_result = client.get_query_results(
                        QueryExecutionId=response_query_execution_id['QueryExecutionId']
                    )
                    result_data = response_query_result['ResultSet']
                    if key == '':
                        if len(result_data['Rows'][1]['Data'][0]) == 0:
                            return 'No Record Found'
                        else:
                            output = result_data['Rows'][1]['Data'][0]['VarCharValue']
                    elif key == 'KPI':
                        output = result_data
                    else:
                        output = status
                    return output
                else:
                    time.sleep(1)
            return False
        except:
            logger.error("Error occured while executing query on Athena", extra=self.execution_context.get_context())
            raise

    def athena_data_validation(self, process_id):
        prev_cnt = []
        curr_cnt = []
        dataset_list = []
        difference = []
        table_list = []
        prev_batch_id_list = []
        current_batch_id_list = []
        status_list = []
        try:
            response_df = self.fetch_dataset_id(process_id)
            env = self.configuration.get_configuration([CommonConstants.ENVIRONMENT_PARAMS_KEY, "environment"])
            query = self.validation_queries[self.validation_queries["query_type"] == "record_cnt_check"].iloc[0][
                'query']
            query_msck = self.validation_queries[self.validation_queries["query_type"] == "refresh_tables"].iloc[0][
                'query']
            status_msg = "The Query to Execute for Athena Count Check with process id " + str(process_id) + ": " + query
            logger.info(status_msg, extra=self.execution_context.get_context())
            for index, row in response_df.iterrows():
                key = 1
                table_name = row["dataset_name"]
                query_msck_final = str(query_msck).replace("$$table_name", str(table_name)).replace('$$env',
                                                                                                    env.lower())
                logger.debug('Query: ' + query_msck_final, extra=self.execution_context.get_context())
                response = self.query_results(query_msck_final, key)
                logger.debug('Refresh status: ' + str(response),
                             extra=self.execution_context.get_context())
            for index, row in response_df.iterrows():
                key = ''
                prev_batch_id = row["prev_batch_id"]
                current_batch_id = row["curr_batch_id"]
                table_name = row["dataset_name"]
                query_prev = str(query).replace("$$table_name", str(table_name)).replace("$$pt_batch_id",
                                                                                         str(prev_batch_id)).replace(
                    '$$env', env.lower())
                status_msg = "The Query to Execute for Athena Previous Count Check for dataset id " + \
                             str(row["dataset_id"]) + ": " + query_prev
                logger.info(status_msg, extra=self.execution_context.get_context())
                query_curr = str(query).replace("$$table_name", str(table_name)).replace("$$pt_batch_id",
                                                                                         str(current_batch_id)).replace(
                    '$$env', env.lower())
                status_msg = "The Query to Execute for Athena Current Count Check for dataset id " + \
                             str(row["dataset_id"]) + ": " + query_curr
                logger.info(status_msg, extra=self.execution_context.get_context())
                response_prev = self.query_results(query_prev, key)
                logger.debug('Output for previous Batch ID: ' + str(response_prev),
                             extra=self.execution_context.get_context())
                response_curr = self.query_results(query_curr, key)
                logger.debug('Output for current Batch ID: ' + str(response_curr),
                             extra=self.execution_context.get_context())
                if str(response_prev) == 'No Record Found' or str(response_curr) == 'No Record Found' or str(
                        response_prev) == 'False' \
                        or str(response_curr) == 'False':
                    cal_difference = float(0)
                elif str(response_prev) == '0' and str(response_curr) \
                        == '0':
                    cal_difference = float(0)
                elif str(response_prev) == '0' and str(response_curr) \
                        != '0':
                    cal_difference = float(100)
                elif str(response_prev) != '0' and str(response_curr) \
                        == '0':
                    cal_difference = float(100)
                else:
                    cal_difference = ((float(response_curr) - float(response_prev)) / float(response_prev)) * 100
                    cal_difference = round(abs(cal_difference), 3)
                # logger.debug('Percentage Difference: ' + str(cal_difference) +'%', extra=self.execution_context.get_context())
                if (abs(cal_difference) > 5):
                    status = 'Fail'
                else:
                    status = 'Pass'
                logger.debug('Status:' + status, extra=self.execution_context.get_context())
                dataset_list.append(str(row["dataset_id"]))
                table_list.append(table_name)
                prev_batch_id_list.append(str(prev_batch_id))
                current_batch_id_list.append(str(current_batch_id))
                prev_cnt.append(response_prev)
                curr_cnt.append(response_curr)
                difference.append(str(cal_difference) + '%')
                status_list.append(str(status))
            list_of_tuples = list(
                zip(dataset_list, table_list, dataset_list, prev_batch_id_list, current_batch_id_list, prev_cnt,
                    curr_cnt, difference, status_list))
            result_set = pd.DataFrame(list_of_tuples, columns=['dataset_id', 'table', 'query', 'previous_ batch_id',
                                                               'current_ batch_id', 'previous_result', 'current_result',
                                                               'difference', 'status'])
            logger.debug(result_set,
                         extra=self.execution_context.get_context())
            # return  result_set
            if len(result_set) == 0:
                process_id_list_str = [str(process_id)]
                response_list = ["No Records Found"]
                list_of_tuples = list(zip(process_id_list_str, response_list))
                res_df = pd.DataFrame(list_of_tuples,
                                      columns=["process_id", "athena_count_check"])
            else:
                res_df = result_set
            return res_df
        except:
            logger.error("Error occured while data validation", extra=self.execution_context.get_context())
            raise

    def ingestion_validation(self, process_id, **kwargs):
        """
        :param process_id:
        :return:
        """
        email_type = "batch_status"
        ses = self.configuration.get_configuration(
            [CommonConstants.ENVIRONMENT_PARAMS_KEY, "ses_region"])
        logger.info("SES Region retrieved from Configurations is: " + str(ses),
                    extra=self.execution_context.get_context())
        email_sender = self.configuration.get_configuration([CommonConstants.ENVIRONMENT_PARAMS_KEY,
                                                             "email_type_configurations",
                                                             str(email_type),
                                                             "ses_sender"])
        logger.info("Email Sender retrieved from Configurations is: " + str(email_sender),
                    extra=self.execution_context.get_context())
        email_recipient_list = self.configuration.get_configuration(
            [CommonConstants.ENVIRONMENT_PARAMS_KEY, "email_type_configurations",
             str(email_type), "ses_recipient_list"])
        logger.info("Email Recipient retrieved from Configurations is: " + str(
            email_recipient_list), extra=self.execution_context.get_context())
        env = self.configuration.get_configuration([CommonConstants.ENVIRONMENT_PARAMS_KEY, "environment"])
        email_subject = "[" + str(env).upper() + "] " + "Ingestion Process | Validation"
        adapter_df = self.adapter_run_time(process_id)
        ingestion_df = self.ingestion_run_time(process_id)
        dqm_df = self.dqm_check(process_id)
        file_count_df = self.file_count_check(process_id)
        time.sleep(120)
        athena_validation_df = self.athena_data_validation(process_id)
        msg = make_ingestion_mail(adapter_df, ingestion_df, file_count_df, dqm_df, athena_validation_df, process_id)
        ses = boto3.client("ses", region_name=self.region)
        print("Email Sender retrieved from Configurations is: " + str(email_sender))
        email_recipient_list = self.configuration.get_configuration(
            [CommonConstants.ENVIRONMENT_PARAMS_KEY, "delta_email_type_configurations", "ses_recipient_list"])
        sender = "sanofi.clinical@zs.com"
        status = self.send_email(sender, email_recipient_list, str(msg), email_subject)
        if status:
            print("Email sent Successfully")
        else:
            print("Check Logs")

    def send_email(self, sender, recipient_list, email_body_string, email_subject):
        try:
            # ses_connection = boto.ses.connect_to_region(ses_region)
            # ses_connection.send_email(source=sender, subject=email_subject,body=email_body_string, to_addresses=recipient_list,format=NotificationUtilityConstants.MIME_TYPE)
            self.configuration = JsonConfigUtility(
                CommonConstants.AIRFLOW_CODE_PATH + '/' + CommonConstants.ENVIRONMENT_CONFIG_FILE)
            msg = MIMEMultipart('alternative')
            msg['From'] = sender
            msg['To'] = ", ".join(recipient_list)
            msg['Subject'] = email_subject
            e_body = MIMEText(email_body_string, 'html')
            msg.attach(e_body)
            smtp_host = self.configuration.get_configuration([CommonConstants.ENVIRONMENT_PARAMS_KEY, "smtp_server"])
            smtp_port = self.configuration.get_configuration([CommonConstants.ENVIRONMENT_PARAMS_KEY, "smtp_port"])
            print("SMTP HOST ", smtp_host, "SMTP PORT ", smtp_port)
            server = smtplib.SMTP(host=smtp_host, port=smtp_port)
            server.connect(smtp_host, smtp_port)
            server.starttls()
            server.send_message(msg)
            server.quit()
            return True
        except:
            raise Exception(
                "Issue Encountered while sending the email to recipients --> " + str(traceback.format_exc()))

    def fetch_pt_cycle_id(self):
        prev_cycle_id = []
        final_process_id = []
        current_cycle_id = []
        prev_count = []
        current_count = []
        try:
            # process_id_list = list(ofr_validation_queries['Process_id'].drop_duplicates())
            logger.debug('Distinct process ids for the queries',
                         extra=self.execution_context.get_context())
            process_id_list = list(self.ofr_validation_queries['Process_id'].drop_duplicates())
            logger.debug(process_id_list,
                         extra=self.execution_context.get_context())
            configuration = JsonConfigUtility(
                os.path.join(CommonConstants.AIRFLOW_CODE_PATH, CommonConstants.ENVIRONMENT_CONFIG_FILE))
            audit_information = configuration.get_configuration(
                [CommonConstants.ENVIRONMENT_PARAMS_KEY, "mysql_db"])
            for process_id in process_id_list:
                logger.debug("Fetching data for process id: " + str(process_id),
                             extra=self.execution_context.get_context())
                sqlquerycurrent = "select max(cycle_id) as curr_cycle_id from " \
                                  + audit_information + ".log_cycle_dtl" \
                                  + " where process_id = " + str(
                    process_id) + " ;"

                logger.debug("Query Output for sqlquerycurrent" + str(sqlquerycurrent),
                             extra=self.execution_context.get_context())

                sqlqueryprev = "select max(cycle_id) as prev_cycle_id from " \
                               + audit_information + ".log_cycle_dtl where cycle_id  !=  (select max(cycle_id) from " \
                               + audit_information + ".log_cycle_dtl where  process_id = " \
                               + str(process_id) + " and cycle_status = " + self.cycle_status_succeed + " ) " \
                               + " and process_id = " + str(
                    process_id) + " ;"
                logger.debug("Query Output for sqlqueryprev" + str(sqlqueryprev),
                             extra=self.execution_context.get_context())
                sqlquery_current_result = MySQLConnectionManager().execute_query_mysql(sqlquerycurrent, False)
                logger.debug("Query Output for current cycle id" + str(sqlquery_current_result),
                             extra=self.execution_context.get_context())
                sqlquery_prev_result = MySQLConnectionManager().execute_query_mysql(sqlqueryprev, False)
                logger.debug("Query Output previous cycle id" + str(sqlquery_prev_result),
                             extra=self.execution_context.get_context())
                final_process_id.append(str(process_id))
                prev_cycle_id.append(str(sqlquery_prev_result[0]['prev_cycle_id']))
                current_cycle_id.append(str(sqlquery_current_result[0]['curr_cycle_id']))
            list_of_tuples = list(zip(final_process_id, prev_cycle_id, current_cycle_id))
            temp_df = pd.DataFrame(list_of_tuples, columns=['process_id', 'prev_cycle_id', 'current_cycle_id'])
            logger.debug("Mapping of Process ID, Previous Cycle ID, Current Cycle ID: ",
                         extra=self.execution_context.get_context())
            logger.debug(temp_df,
                         extra=self.execution_context.get_context())
            return temp_df

        except:
            logger.error("Error occured while data validation", extra=self.execution_context.get_context())
            raise

    def dw_dedupe_data_validation(self, process_id, **kwargs):
        query_list = []
        curr_cnt = []
        category_list = []
        logger.debug("Output for ofr_validation_queries df ",
                     extra=self.execution_context.get_context())
        logger.debug(self.ofr_validation_queries,
                     extra=self.execution_context.get_context())
        try:
            mapping_df = self.fetch_pt_cycle_id()
            env = self.configuration.get_configuration(
                [CommonConstants.ENVIRONMENT_PARAMS_KEY, "env"])
            for index, row in (self.ofr_validation_queries.iterrows()):
                key = ''
                if row['Category'] == self.msck_repair_tc_key:
                    key = 1
                    query = row['Queries'].replace('$$db_env',
                                                   env.lower()).replace(
                        '$$env', env.lower())
                    response = self.query_results(query, key)
                    logger.debug('Test Case: ' + row['Category'],
                                 extra=self.execution_context.get_context())
                    logger.debug('Query: ' + query,
                                 extra=self.execution_context.get_context())
                    logger.debug('Refresh status: ' + str(response),
                                 extra=self.execution_context.get_context())
                elif str(row['Process_id']) == str(process_id):
                    prev_cycle_id = mapping_df[mapping_df['process_id'] == str(
                        row['Process_id'])].iloc[0][
                        'prev_cycle_id']
                    current_cycle_id = mapping_df[
                        mapping_df['process_id'] == str(
                            row['Process_id'])].iloc[0][
                        'current_cycle_id']
                    query_curr = row['Queries'].replace('$$db_env',
                                                        env.lower()).replace(
                        '$$pt_cycle_id$$',
                        str(current_cycle_id)).replace(
                        '$$env', env.lower())
                    logger.debug('query_curr: ' + query_curr,
                                 extra=self.execution_context.get_context())
                    response_curr = self.query_results(query_curr, key)
                    logger.debug('Test Case: ' + row['Category'],
                                 extra=self.execution_context.get_context())
                    logger.debug('Query: ' + query_curr,
                                 extra=self.execution_context.get_context())
                    logger.debug(
                        'Output for current cycle ID: ' + str(response_curr),
                        extra=self.execution_context.get_context())
                    category_list.append(row['Category'])
                    query_list.append(query_curr)
                    curr_cnt.append(response_curr)
            result_set = pd.DataFrame(
                {'Query': query_list, 'Category': category_list, 'Current Result': curr_cnt})
            result_set1 = result_set[
                result_set['Category'] != self.msck_repair_tc_key]
            logger.info("printing resultset")
            logger.debug(result_set1,
                         extra=self.execution_context.get_context())
            ofr_dw_email_type = 'cycle_status'
            dag_name = 'Sanofi-DataEngineering-L2-Fact-A1-Reporting-Monthly'
            email_recipient_list = self.configuration.get_configuration(
                [CommonConstants.ENVIRONMENT_PARAMS_KEY,
                 "delta_email_type_configurations", "ses_recipient_list"])
            sender = "sanofi.clinical@zs.com"
            email_subject = "DEDUPE VALIDATION"
            msg = make_dw_validation_mail(result_set)
            status = self.send_email(sender, email_recipient_list, str(msg),
                                     email_subject)
            if status:
                print("Email sent Successfully")
            else:
                print("Check Logs")
            output_file_name = self.ofr_validation_output_file_name + '.csv'
            output_file_path = self.ofr_validation_output_file_path
            file_ec2_path = output_file_name
            result_set1.to_csv(file_ec2_path, encoding='utf-8', index=False)
            s3_file_name = self.ofr_validation_output_file_name + '_' + str(
                datetime.today().strftime("%Y%m%d%H%M%S")) + '.csv'
            s3_sandbox_bucket = self.configuration.get_configuration(
                [CommonConstants.ENVIRONMENT_PARAMS_KEY, "s3_sandbox_bucket"])
            file_s3_path = self.ofr_validation_output_file_s3_path + s3_file_name
            self.s3.upload_file(file_ec2_path, s3_sandbox_bucket, file_s3_path)
            logger.debug(
                "Output file has been uploaded on S3: " + s3_file_name,
                extra=self.execution_context.get_context())
        except:
            logger.error("Error occured while data validation", extra=self.execution_context.get_context())
            raise

    def dw_data_validation(self, process_id, **kwargs):
        prev_cnt = []
        curr_cnt = []
        query_list = []
        difference = []
        category_list = []
        layer_list = []
        table_list = []
        prev_cycle_id_list = []
        current_cycle_id_list = []
        status_list = []
        # logger.debug("Output for ofr_validation_queries " + self.ofr_validation_queries,extra=self.execution_context.get_context())
        logger.debug("Output for ofr_validation_queries df ",
                     extra=self.execution_context.get_context())
        logger.debug(self.ofr_validation_queries,
                     extra=self.execution_context.get_context())
        try:
            mapping_df = self.fetch_pt_cycle_id()
            logger.info("Output for mapping_df " + mapping_df)
            logger.info("process id dag id " + str(process_id))
            env = self.configuration.get_configuration([CommonConstants.ENVIRONMENT_PARAMS_KEY, "env"])
            # ofr_validation_queries_df = self.ofr_validation_queries(self.ofr_validation_queries)
            for index, row in (self.ofr_validation_queries.iterrows()):
                logger.info('process id from dataframe ' + str(row['Process_id']))
                key = ''
                if row['Category'] == self.msck_repair_tc_key:
                    key = 1
                    query = row['Queries'].replace('$$db_env', env.lower()).replace('$$env', env.lower())
                    logger.debug('Test Case: ' + row['Category'],
                                 extra=self.execution_context.get_context())
                    logger.debug('Query: ' + query, extra=self.execution_context.get_context())
                    response = self.query_results(query, key)
                    logger.debug('Refresh status: ' + str(response),
                                 extra=self.execution_context.get_context())
                elif str(row['Process_id']) == str(process_id):
                    logger.info('in elif block')
                    prev_cycle_id = mapping_df[mapping_df['process_id'] == str(row['Process_id'])].iloc[0][
                        'prev_cycle_id']
                    logger.debug('previous cycle id: ' + prev_cycle_id,
                                 extra=self.execution_context.get_context())
                    current_cycle_id = mapping_df[mapping_df['process_id'] == str(row['Process_id'])].iloc[0][
                        'current_cycle_id']
                    logger.debug('current cycle id: ' + prev_cycle_id,
                                 extra=self.execution_context.get_context())
                    query_prev = row['Queries'].replace('$$db_env', env.lower()).replace('$$pt_cycle_id$$',
                                                                                         str(prev_cycle_id)).replace(
                        '$$env', env.lower())
                    logger.debug('Test Case: ' + row['Category'],
                                 extra=self.execution_context.get_context())
                    logger.debug('Query: ' + query_prev, extra=self.execution_context.get_context())
                    response_prev = self.query_results(query_prev, key)
                    logger.debug('Output for previous cycle ID: ' + str(response_prev),
                                 extra=self.execution_context.get_context())
                    # query_curr = row['Queries'].replace('$$db_env', env.lower()).replace('$$pt_cycle_id$$',
                    ##str(current_cycle_id)).replace(
                    ##    '$$env', env.lower())
                    query_curr = row['Queries'].replace('$$db_env', env.lower()).replace('$$pt_cycle_id$$',
                                                                                         str(current_cycle_id)).replace(
                        '$$env', env.lower()).replace('$$current_cycle_id$$', str(current_cycle_id)).replace(
                        '$$prev_cycle_id$$', str(prev_cycle_id))
                    logger.debug('Test Case: ' + row['Category'],
                                 extra=self.execution_context.get_context())
                    logger.debug('Query: ' + query_curr, extra=self.execution_context.get_context())
                    response_curr = self.query_results(query_curr, key)
                    logger.debug('Output for current cycle ID: ' + str(response_curr),
                                 extra=self.execution_context.get_context())
                    if str(response_prev) == 'No Record Found' or str(response_curr) == 'No Record Found' or str(
                            response_prev) == 'False' \
                            or str(response_curr) == 'False':
                        cal_difference = float(0)
                    elif str(response_prev) == '0.0' and str(response_curr) \
                            == '0.0':
                        cal_difference = float(0)
                    elif str(response_prev) == '0.0' and str(response_curr) \
                            != '0.0':
                        cal_difference = float(100)
                    elif str(response_prev) != '0.0' and str(response_curr) \
                            == '0.0':
                        cal_difference = float(100)
                    elif str(response_prev) == '0' and str(response_curr) \
                            == '0':
                        cal_difference = float(0)
                    elif str(response_prev) == '0' and str(response_curr) \
                            != '0':
                        cal_difference = float(100)
                    elif str(response_prev) != '0' and str(response_curr) \
                            == '0':
                        cal_difference = float(100)
                    else:
                        cal_difference = ((float(response_curr) - float(response_prev)) / float(response_prev)) * 100
                        cal_difference = round(abs(cal_difference), 3)
                    logger.debug('Percentage Difference: ' + str(cal_difference) + '%',
                                 extra=self.execution_context.get_context())
                    if (abs(cal_difference) > 5):
                        status = 'Fail'
                    else:
                        status = 'Pass'
                    logger.debug('Status:' + status, extra=self.execution_context.get_context())
                    category_list.append(row['Category'])
                    layer_list.append(row['Layer'])
                    query_list.append(row['Queries'])
                    table_list.append(row['Table'])
                    prev_cycle_id_list.append(prev_cycle_id)
                    current_cycle_id_list.append(current_cycle_id)
                    if row['Category'] in ["Duplicate Count", "Retention Queries"]:
                        prev_cnt.append("NA")
                        status_list.append("NA")
                        difference.append("NA")
                    else:
                        prev_cnt.append(response_prev)
                        status_list.append(status)
                    curr_cnt.append(response_curr)
                    difference.append(str(cal_difference) + '%')

            list_of_tuples = list(
                zip(category_list, layer_list, table_list, prev_cycle_id_list, current_cycle_id_list,
                    prev_cnt,
                    curr_cnt, difference, status_list))
            result_set = pd.DataFrame(list_of_tuples,
                                      columns=['Category', 'Layer', 'Table', 'Prev Cycle ID', 'Curr Cycle ID',
                                               'Previous Result', 'Current Result', 'Difference', 'Status'])
            result_set1 = result_set[result_set['Category'] != self.msck_repair_tc_key]
            logger.debug(result_set1,
                         extra=self.execution_context.get_context())
            ofr_dw_email_type = 'cycle_status'
            dag_name = 'Sanofi-DataEngineering-L2-Fact-A1-Reporting-Monthly'
            # DagUtils.trigger_notification_utility(ofr_dw_email_type, process_id=3000, dag_name=dag_name,
            #                                      emr_status=CommonConstants.STATUS_SUCCEEDED,
            #                                      final_result=result_set1)
            email_recipient_list = self.configuration.get_configuration(
                [CommonConstants.ENVIRONMENT_PARAMS_KEY, "delta_email_type_configurations", "ses_recipient_list"])
            sender = "sanofi.clinical@zs.com"
            if str(process_id) in ["3000"]:
                email_subject = "REPORTING VALIDATION"
            else:
                email_subject = "POSTDEDUPE VALIDATION"
            msg = make_dw_validation_mail(result_set)
            status = self.send_email(sender, email_recipient_list, str(msg), email_subject)
            if status:
                print("Email sent Successfully")
            else:
                print("Check Logs")
            output_file_name = self.ofr_validation_output_file_name + '.csv'
            output_file_path = self.ofr_validation_output_file_path
            file_ec2_path = output_file_name
            result_set1.to_csv(file_ec2_path, encoding='utf-8', index=False)
            s3_file_name = self.ofr_validation_output_file_name + '_' + str(
                datetime.today().strftime("%Y%m%d%H%M%S")) + '.csv'
            s3_sandbox_bucket = self.configuration.get_configuration(
                [CommonConstants.ENVIRONMENT_PARAMS_KEY, "s3_sandbox_bucket"])
            file_s3_path = self.ofr_validation_output_file_s3_path + s3_file_name
            self.s3.upload_file(file_ec2_path, s3_sandbox_bucket, file_s3_path)
            logger.debug("Output file has been uploaded on S3: " + s3_file_name,
                         extra=self.execution_context.get_context())

        except:
            logger.error("Error occured while data validation", extra=self.execution_context.get_context())
            raise

    def update_run_time(self, process_id, data_date, run_time):
        try:
            if os.path.exists(self.ec2_excel_path):
                os.remove(self.ec2_excel_path)
            else:
                pass
            s3 = boto3.resource('s3', region_name=self.region)
            result = s3.meta.client.download_file(self.s3_sandbox_bucket, self.run_time_path, self.ec2_excel_path)
            response = str(data_date)[:7]
            run_time_add_obj = self.s3.get_object(Bucket=self.s3_sandbox_bucket, Key=self.run_time_path)
            run_time_add = pd.read_excel(io.BytesIO(run_time_add_obj['Body'].read()), None)
            result = list(run_time_add.keys())
            count = 0
            create_sheet_object = self.s3.get_object(Bucket=self.s3_sandbox_bucket, Key=self.run_time_path)
            for res in result:
                if res == response:
                    df = pd.read_excel(io.BytesIO(create_sheet_object['Body'].read()), str(res))
                    df_col_list = list(df.columns)
                    col_count = 0
                    for col in df_col_list:
                        if col == str(data_date):
                            col_count = col_count + 1
                        else:
                            pass
                    if col_count == 0:
                        df[str(data_date)] = ""
                        for i, r in df.iterrows():
                            if str(int(r["process_id"])) == str(process_id) and (
                                    str(r["process_type"]).lower() == "ingestion" or
                                    str(r["process_type"]).lower() == "dw"):
                                update_df = pd.DataFrame({str(data_date): [str(run_time)]}, index=[i])
                                df.update(update_df)
                                self.write_excel_file(response, df)
                                self.s3.upload_file(self.ec2_excel_path, self.s3_sandbox_bucket, self.run_time_path)
                            else:
                                pass
                    else:
                        for i, r in df.iterrows():
                            if str(int(r["process_id"])) == str(process_id) and (
                                    str(r["process_type"]).lower() == "ingestion" or
                                    str(r["process_type"]).lower() == "dw"):
                                update_df = pd.DataFrame({str(data_date): [str(run_time)]}, index=[i])
                                df.update(update_df)
                                self.write_excel_file(response, df)
                                self.s3.upload_file(self.ec2_excel_path, self.s3_sandbox_bucket, self.run_time_path)
                            else:
                                pass
                    count = count + 1
                else:
                    pass
            if count == 0:
                response = str(data_date).split("-")
                new_sheet_name = str((response[0])) + "-" + response[1]
                self.run_time_df[str(data_date)] = ""
                new_df = self.run_time_df
                for i, r in new_df.iterrows():
                    if str(int(r["process_id"])) == str(process_id) and (
                            str(r["process_type"]).lower() == "ingestion" or
                            str(r["process_type"]).lower() == "dw"):
                        update_df = pd.DataFrame({str(data_date): [str(run_time)]}, index=[i])
                        new_df.update(update_df)
                        self.write_excel_file(new_sheet_name, new_df)
                        self.s3.upload_file(self.ec2_excel_path, self.s3_sandbox_bucket, self.run_time_path)
                    else:
                        pass
            else:
                pass
        except Exception as e:
            error = "ERROR in " + "Update Monthly Run Tracker" + \
                    "For Proceess ID : " + str(process_id) + \
                    " ERROR MESSAGE: " + str(e)
            self.execution_context.set_context({"traceback": error})
            logger.error(error, extra=self.execution_context.get_context())

    def adapter_update_run_time(self, process_name, data_date, run_time):
        try:
            if os.path.exists(self.ec2_excel_path):
                os.remove(self.ec2_excel_path)
            else:
                pass
            s3 = boto3.resource('s3', region_name=self.region)
            result = s3.meta.client.download_file(self.s3_sandbox_bucket, self.run_time_path, self.ec2_excel_path)
            response = str(data_date)[:7]
            run_time_add_obj = self.s3.get_object(Bucket=self.s3_sandbox_bucket, Key=self.run_time_path)
            run_time_add = pd.read_excel(io.BytesIO(run_time_add_obj['Body'].read()), None)
            result = list(run_time_add.keys())
            count = 0
            create_sheet_object = self.s3.get_object(Bucket=self.s3_sandbox_bucket, Key=self.run_time_path)
            for res in result:
                if res == response:
                    df = pd.read_excel(io.BytesIO(create_sheet_object['Body'].read()), str(res))
                    df_col_list = list(df.columns)
                    col_count = 0
                    for col in df_col_list:
                        if col == str(data_date):
                            col_count = col_count + 1
                        else:
                            pass
                    if col_count == 0:
                        df[str(data_date)] = ""
                        for i, r in df.iterrows():
                            if str(r["process_name"]) == str(process_name) and (
                                    str(r["process_type"]).lower() == "adapter"):
                                update_df = pd.DataFrame({str(data_date): [str(run_time)]}, index=[i])
                                df.update(update_df)
                                self.write_excel_file(response, df)
                                self.s3.upload_file(self.ec2_excel_path, self.s3_sandbox_bucket, self.run_time_path)
                            else:
                                pass
                    else:
                        for i, r in df.iterrows():
                            if str(r["process_name"]) == str(process_name) and (
                                    str(r["process_type"]).lower() == "adapter"):
                                update_df = pd.DataFrame({str(data_date): [str(run_time)]}, index=[i])
                                df.update(update_df)
                                self.write_excel_file(response, df)
                                self.s3.upload_file(self.ec2_excel_path, self.s3_sandbox_bucket, self.run_time_path)
                            else:
                                pass
                    count = count + 1
                else:
                    pass
            if count == 0:
                response = str(data_date).split("-")
                new_sheet_name = str((response[0])) + "-" + response[1]
                self.run_time_df[str(data_date)] = ""
                new_df = self.run_time_df
                for i, r in new_df.iterrows():
                    if str(r["process_name"]) == str(process_name) and (str(r["process_type"]).lower() == "adapter"):
                        update_df = pd.DataFrame({str(data_date): [str(run_time)]}, index=[i])
                        new_df.update(update_df)
                        self.write_excel_file(new_sheet_name, new_df)
                        self.s3.upload_file(self.ec2_excel_path, self.s3_sandbox_bucket, self.run_time_path)
                    else:
                        pass
            else:
                pass
        except Exception as e:
            error = "ERROR in " + "Adapter Monthly update Run Time" + \
                    "For Proceess name : " + str(process_name) + \
                    " ERROR MESSAGE: " + str(e)
            self.execution_context.set_context({"traceback": error})
            logger.error(error, extra=self.execution_context.get_context())

    def write_excel_file(self, sheet_name, new_df):
        try:
            writer = pd.ExcelWriter(self.ec2_excel_path, engine='openpyxl')
            writer.book = load_workbook(self.ec2_excel_path)
            writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
            new_df.to_excel(writer, index=False, sheet_name=str(sheet_name))
            writer.save()
        except Exception as e:
            error = "ERROR in " + "Writing to Excel file" + str(e)
            self.execution_context.set_context({"traceback": error})
            logger.error(error, extra=self.execution_context.get_context())

    def kpi_ops(self, **kwargs):
        source_list1 = []
        bacth_id_list1 = []
        counts_list1 = []
        try:
            # cmd = "python Glue_Crawler.py -c aws-a0199-use1-00-p-gcr-sanofi-ctf-ds-staging"
            # execution_status = CommonUtils().execute_shell_command(cmd)
            # if execution_status != True:
            #    raise Exception("Failed To Kinit with provided kerberos details")
            key = 'KPI'
            kpi_query1 = """select * from (
            select * from (select distinct 'citeline_trial' as source, pt_batch_id, count(distinct trial_id) as count from sanofi_ctfo_datastore_staging_prod.citeline_trialtrove group by 2 order by pt_batch_id desc limit 2)
            UNION ALL
            select * from (select distinct 'citeline_sites' as source, pt_batch_id, count(distinct site_id) as count from sanofi_ctfo_datastore_staging_prod.citeline_organization group by 2 order by pt_batch_id desc limit 2)
            UNION ALL
            select * from (select distinct 'citeline_investigator' as source, pt_batch_id, count(distinct investigator_id) as count from sanofi_ctfo_datastore_staging_prod.citeline_investigator group by 2 order by pt_batch_id desc limit 2)
            UNION ALL
            select * from (select distinct 'aact_trial' as source, pt_batch_id, count(distinct nct_id) as count from sanofi_ctfo_datastore_staging_prod.aact_studies group by 2 order by pt_batch_id desc limit 2)
            UNION ALL
            select * from (select distinct 'aact_sites' as source, pt_batch_id, count(distinct id) as count from sanofi_ctfo_datastore_staging_prod.aact_facilities group by 2 order by pt_batch_id desc limit 2)
            UNION ALL
            select * from (select distinct 'aact_investigator' as source, pt_batch_id, count(distinct id) as count from sanofi_ctfo_datastore_staging_prod.aact_facility_investigators group by 2 order by pt_batch_id desc limit 2)
            UNION ALL
            select * from (select distinct 'dqs_trial' as source, pt_batch_id, count(distinct member_study_id) as count from sanofi_ctfo_datastore_staging_prod.dqs_study group by 2 order by pt_batch_id desc limit 2)
            UNION ALL
            select * from (select distinct 'dqs_sites' as source, pt_batch_id, count(distinct facility_golden_id) as count from sanofi_ctfo_datastore_staging_prod.dqs_study group by 2 order by pt_batch_id desc limit 2)
           ) order by source, pt_batch_id desc"""
            kpi_query2 = """select * from (
                    select * from (select distinct 'dqs_investigator' as source, pt_batch_id, count(distinct person_golden_id) as count from sanofi_ctfo_datastore_staging_prod.dqs_study group by 2 order by pt_batch_id desc limit 2)
            UNION ALL
            select * from (select distinct 'ctms_trial' as source, pt_batch_id, count(distinct study_code) as count from sanofi_ctfo_datastore_staging_prod.ctms_study_site group by 2 order by pt_batch_id desc limit 2)
            UNION ALL
            select * from (select distinct 'ctms_sites' as source, pt_batch_id, count(distinct center_id) as count from sanofi_ctfo_datastore_staging_prod.ctms_study_site group by 2 order by pt_batch_id desc limit 2)
            UNION ALL
            select * from (select distinct 'ctms_investigator' as source, pt_batch_id, count(distinct primary_investigator_id) as count from sanofi_ctfo_datastore_staging_prod.ctms_study_site group by 2 order by pt_batch_id desc limit 2)
           ) order by source, pt_batch_id desc"""
            athena_json = self.query_results(kpi_query1, key)
            athena_json2 = self.query_results(kpi_query2, key)
            for each_row in range(1, len(athena_json['Rows'])):
                source_list1.append(str(athena_json['Rows'][each_row]['Data'][0]['VarCharValue']))
                bacth_id_list1.append(str(athena_json['Rows'][each_row]['Data'][1]['VarCharValue']))
                counts_list1.append(str(athena_json['Rows'][each_row]['Data'][2]['VarCharValue']))
            for each_row in range(1, len(athena_json2['Rows'])):
                source_list1.append(str(athena_json2['Rows'][each_row]['Data'][0]['VarCharValue']))
                bacth_id_list1.append(str(athena_json2['Rows'][each_row]['Data'][1]['VarCharValue']))
                counts_list1.append(str(athena_json2['Rows'][each_row]['Data'][2]['VarCharValue']))
            list_of_tuples = list(zip(source_list1, bacth_id_list1, counts_list1))
            result_set = pd.DataFrame(list_of_tuples, columns=['Source', 'Batch_ID', 'Count'])
            email_recipient_list = ["dhananjay.nagare@zs.com", "vivek.pangasa@zs.com", "richa.d.singh@zs.com"]
            sender = "sanofi.clinical@zs.com"
            email_subject = "Sanofi - Weekly OPS KPIs"
            msg = make_kpi_ops_mail(result_set)
            status = self.send_email(sender, email_recipient_list, str(msg), email_subject)
            if status:
                print("Email sent Successfully")
            else:
                print("Check Logs")
        except Exception as e:
            error = "ERROR occurred while performing KPI checks" + str(e)
            self.execution_context.set_context({"traceback": error})
            logger.error(error, extra=self.execution_context.get_context())


if __name__ == '__main__':
    process_id = sys.argv[1]
    new_obj = OFRDataValidationUtility()
    if process_id == "DW":
        new_obj.dw_data_validation()
    elif process_id == "diseaseQC":
        new_obj.disease_mapping_qc(process_id)
    elif process_id == "taQC":
        new_obj.ta_mapping_qc(process_id)
    elif process_id == "KPI":
        new_obj.kpi_ops()
    else:
        new_obj.ingestion_validation(process_id)
